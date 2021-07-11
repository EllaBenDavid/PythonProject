##Ella Ben David 206821332
##Diana Medvedovsky 316080456

import pyodbc  #SQL
import openpyxl  #EXCEL
from tkinter import *  #MAKE WINDOW
from tkinter import messagebox  #MESSAGEBOX - INFO
from PIL import ImageTk,Image  #LOGO FOR THE FIRST WINDOW
import requests  #WEATHER
import json  #WEATHER
import time as tm #CURRENT_HOUR


##_______CLASSES_______##
class Customer:
    def __init__(self,ID,name,gender,money):
        self.ID=ID
        self.name=name
        self.gender=gender
        self.money=int(money)
        self.myProducts=[]
        self.quantity=0
        
    def BuyProduct(self,p):
        print("please choose a Product below: ")
        p.Print()
        choice = int(input())
        if choice>(len(p.products)):
            print('There is no such product')
            return -1
        else: 
            self.myProducts.append(p.products[choice-1]) 
            self.quantity += 1
            self.money -= p.products[choice-1].price
            print("money left {}$".format(self.money))
            return choice-1
            
            
    
    def ShowMyProducts(self):
        if len(self.myProducts)>0:
            print("{} Products are these :".format(self.name))
            for i in self.myProducts:
                print("*************************************")
                print("Product code {} \nthe product name is {} \nand the product price is: {}".format(i.code,i.name,i.price))
                print("*************************************") 
        else: print('{} GO TO DO SOME SHOPPING ## '.format(self.name))
            
class Product:
    def __init__(self,code,name,price):
        self.code=code
        self.name=name
        self.price=price

class Store:
    def __init__(self,products):
        self.products = products
    
    def ProductSold(self,choice,name):
        print("the product {} has been sold to {}".format(self.products[choice].name, name))
        self.products.remove(self.products[choice])
    
    def Print(self):
        print("Products left in the store:")
        i=1
        for product in self.products:
            print("{} -{}  Price - {}$ code :{}".format(i,product.name,product.price,product.code))
            i += 1 
            
    def ChangeProduct(self,productIndex,custo,p):
        moneyBeforeChange=custo.money
        if productIndex < len(custo.myProducts):
            returnProd = custo.myProducts[productIndex]
            self.products.append(custo.myProducts[productIndex])
            print("{} bring back {} - {}".format(custo.name,returnProd.name,returnProd.code))
            custo.myProducts.pop(productIndex)
            custo.quantity-=1
            #-----
            print('--- pick a new product --- ')
            choice=custo.BuyProduct(p)
            if choice!=-1:
                p.ProductSold(choice,custo.name)
                custo.money += returnProd.price
                moneyAfterChange=custo.money
                if moneyBeforeChange > moneyAfterChange:
                    left=moneyBeforeChange-moneyAfterChange
                    print("{} need to add {} more ".format(custo.name,left))
                if moneyAfterChange > moneyBeforeChange:
                    left=moneyAfterChange-moneyBeforeChange
                    print("{} got {} back ".format(custo.name,left))
            else: 
                print('pick other...')
        else: print('you dont have such product')

##_______SQL_______##
#שמירת נתונים 
def ExcNoneQuery(sql_command, values):
    server= 'LAPTOP-SM2OPB1U\SQLEXPRESS'
    database = 'pythonProject'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; \
                        SERVER=' + server +'; \
                        DATABASE='+ database +'; \
                        Trusted_connection=yes;')
    crsr = cnxn.cursor()
    count = crsr.execute(sql_command, values).rowcount
    print(str(count) + ' rows affected!')
    crsr.commit()
    crsr.close()
    cnxn.close()

#הדפסת נתונים
def ExcQuery(sql_command, values):
    server= 'LAPTOP-SM2OPB1U\SQLEXPRESS'
    database = 'pythonProject'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; \
                        SERVER=' + server +'; \
                        DATABASE='+ database +'; \
                        Trusted_connection=yes;')

    crsr = cnxn.cursor()
    crsr.execute(sql_command, values )
    
    
    for row in crsr:              
        print(row)
    
    maxExpensiveinQuery('select * from myProductsTB Where price > ? and price = (select max(price) from myProductsTB) order by price desc',(0))
    crsr.close()
    cnxn.close()


def maxExpensiveinQuery(sql_command,values):
    server= 'LAPTOP-SM2OPB1U\SQLEXPRESS'
    database = 'pythonProject'

    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; \
                        SERVER=' + server +'; \
                        DATABASE='+ database +'; \
                        Trusted_connection=yes;')
    crsr = cnxn.cursor()
    crsr.execute(sql_command, values)
    table=[]
    for row in crsr:
        table.append(row)
        print('the most expensive product is '+row[0]+' cost - '+str(row[1])+'$')    
    
    crsr.close()
    cnxn.close()

##_______EXCEL_______##
def saveExcel(wbName,sheetName,p):
    wb = openpyxl.load_workbook(wbName)
    sheet = wb[sheetName]
    #מחיקת הנתונים הקיימים בגליון
    for row in sheet['A1' : sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate]:
        for cell in row:
            cell.value = None
            
    rows=len(p.products)
    cols=3
    start_row_index=2
    start_col_index=1

    #create the table values
    count=0
    for i in range(1,rows+1):
        for j in range(0,cols-(cols-start_col_index)): 
            sheet.cell(row=i, column=j+1).value = p.products[i-1].code 
            count+=1

    for i in range(1,rows+1):
        for j in range(1,cols-(cols-(start_col_index+1))):
            sheet.cell(row=i, column=j+1).value = p.products[i-1].name 
            count+=1
            
    for i in range(1,rows+1):
        for j in range(2,cols-(cols-(start_col_index+2))):
            sheet.cell(row=i, column=j+1).value = p.products[i-1].price 
            count+=1
            
    sheet.cell(start_row_index-1, start_col_index).value = 'Code'
    sheet.cell(start_row_index-1, start_col_index+1).value = 'Name'        
    sheet.cell(start_row_index-1, start_col_index+2).value = 'Price'          
    
    wb.save(wbName)
    print(wbName+' created with '+str(count)+' rows')
    
def readExcel(wbName,sheetName):
    wb = openpyxl.load_workbook(wbName)
    sheet = wb[sheetName]
    rows=len(p.products)
    cols=3
    start_row_index=2
    start_col_index=1
    for i in range(1,rows+1):
        for j in range(0,cols-(cols-start_col_index)): 
            print(sheet.cell(row=i, column=j+1).value) 
    
    for i in range(1,rows+1):
        for j in range(1,cols-(cols-(start_col_index+1))):
            print(sheet.cell(row=i, column=j+1).value)

    for i in range(1,rows+1):
        for j in range(2,cols-(cols-(start_col_index+2))):
            print(sheet.cell(row=i, column=j+1).value) 

##_______BTN_______##
products=[Product('001','swimsuit',120),Product('002','belt',80),
        Product('003','pants',200),Product('004','skirt',100),
        Product('005','sneakers',250),Product('006','coat',400),
        Product('007','dress',130),Product('008','underpants',30),
        Product('009','slippers',45),Product('010','hat',80),
        Product('011','socks',50),Product('012','bag',60),
        Product('013','sunglasse',100),Product('014','pajamas',230),
        Product('015','sweater',160),Product('016','jeans',260),
        Product('017','suit',550)]

customer=Customer(0,'','',0)

p=Store(products)

def btn_print():
    p.Print()
    
def btn_BuyProduct():
    global customer
    choice=customer.BuyProduct(p)
    if choice == -1:
        print('')
    else:p.ProductSold(choice,customer.name)
    
def btn_ChangeProduct():
    global p
    global customer
    if len(customer.myProducts) > 0:
        print("choose the product you want to return")
        j=1
        for i in customer.myProducts:
            print("{} -{}  code :{} Price - {}$".format(j,i.name,i.code,i.price))            
            j+=1
        index = int(input())    
        p.ChangeProduct(index-1,customer,p) 
    else:
        print('sorry you dont have any products to bring back :( ')
        
def btn_GetReceipt():
    if len(customer.myProducts) > 0:
        f = open(r'C:\Users\ellab\OneDrive\שולחן העבודה\Receipt.txt','r+')
        f.write("Hey {}, its your oreder details: \nYou by {} products:\n".format(customer.name,customer.quantity))
        j=1
        sum=0
        for i in customer.myProducts:
            f.write("{}- {} code :{} Price - {}$\n ".format(j,i.name,i.code,i.price))
            j+=1
            sum+=i.price
        f.write('SUM ORDER : {} \n+TAX( 17% ) : {}\n^ TOTAL : {} ^\n**THANKS FOR BUYING IN OUR STORE**\n'.format(sum,round(sum*0.17,2),round(sum*1.17,2)))
        f.close()
        print("**THANK'S FOR BUYING IN OUR STORE**")
    else:
        print('{} GO TO DO SOME SHOPPING ## '.format(customer.name))
    
def btn_viewMyProduct():
    customer.ShowMyProducts()

def btn_Weather_Now(cityBox,WindowWeather):        
        api_request = requests.get("http://api.openweathermap.org/data/2.5/weather?q="+cityBox+"&appid=90978e389f0ca4f4ebec4302716e86da")
        api = json.loads(api_request.content)
        city = api['name']
        country = api['sys']['country']
        discrip_weather = api['weather'][0]['description']
        current_temp = int(api['main']['temp']-273.15)
        Min_temp = int( api['main']['temp_min']- 273.15)
        Max_temp = int( api['main']['temp_max']- 273.15)
        #Creating Labels as description
        city_label = Label(WindowWeather, text="City: "+city, background='pink',fg='VioletRed3',font='bold')
        city_label.grid(row=1,column=2)   
        country_label = Label(WindowWeather, text="country: " + country,bg='pink',fg='VioletRed3',font='bold')
        country_label.grid( row=2, column=2 )
        discrip_weather_label = Label( WindowWeather, text="weather: " + discrip_weather,bg='pink',fg='VioletRed3',font='bold')
        discrip_weather_label.grid( row=3, column=2)
        Min_temp_label = Label( WindowWeather, text="Min Temp: " + str(Min_temp),bg='pink',fg='VioletRed3',font='bold')
        Min_temp_label.grid( row=1, column=3,)
        Max_temp_label = Label( WindowWeather, text="Max Temp: " + str(Max_temp),bg='pink',fg='VioletRed3',font='bold')
        Max_temp_label.grid( row=2, column=3 )
        current_temp_label = Label( WindowWeather, text="current: " + str(current_temp), bg='pink',fg='VioletRed3' ,font='bold')
        current_temp_label.grid( row=4, column=2,)

def btn_Current_Weather(): #find current weather in spesific city in the world
    WindowWeather = Tk()
    WindowWeather.title('Weather App')
    WindowWeather.config(bg='pink')
    WindowWeather.geometry("670x140")
    LabelW = Label(WindowWeather, text ="Where Do You Live?",font = ("Helvetica 16 bold italic",18),bg = "pink",fg='VioletRed3')
    LabelW.grid(row=0, column=0)

    #Create Entry Box
    city_box = Entry(WindowWeather,width=20)
    city_box.grid(row=1, column=1)

    #Create Text Box Lables
    city_label = Label(WindowWeather, text="City",font = ("Helvetica 16 bold italic",15),bg = "pink",fg='VioletRed3')
    city_label.grid(row=1, column=0)

    #Create Button
    city_button = Button(WindowWeather, text="Weather Now", command= lambda: btn_Weather_Now(city_box.get(),WindowWeather))
    city_button.grid(row=2,column=0)

    WindowWeather.mainloop()

def btn_current_Time(window2):
        current_time=tm.strftime('%H:%M:%S')   
        Label3=Label(window2, text=current_time ,font = ("Helvetica 16 italic",15),bg = "pink",fg='VioletRed3')
        Label3.grid(column=4,row=350)

def btn_SaveDB():
    for i in customer.myProducts:
        sql_command = 'Insert into myProductsTB(name,price) VALUES(?,?)'
        values = (i.name, i.price)
        ExcNoneQuery(sql_command, values)
    print(str(len(customer.myProducts))+' Products inserted to db')

def btn_PrintDB():
    sql_command = 'Select * from myProductsTB Where price > ? Order By name'
    values = (0)
    ExcQuery(sql_command, values)
    
def btn_SaveExcel():
    saveExcel(r'C:\Users\ellab\OneDrive\שולחן העבודה\productsList.xlsx','products',p)
    
def btn_PrintExcel():
    readExcel(r'C:\Users\ellab\OneDrive\שולחן העבודה\productsList.xlsx','products')
    
def btn_PrintTxtFile():
    f=open(r'C:\Users\ellab\OneDrive\שולחן העבודה\productsList.txt')
    content= f.read()
    print(content)
    f.close()

##_______GUI_______##

name=''
money=0
Id=''
def SubmitMsg(window1,Id,name,money):
    if customer.gender==0:
        gender="Male"
    else:
        gender="Female" 
    customer.ID= Id
    customer.name=name
    customer.gender=gender
    customer.money=int(money)
    messagebox.showinfo("Setup", "Hello "+customer.name+", ID: "+str(customer.ID) +' | Gender: '+gender+'\nBudget: '+str(customer.money))
    window1.destroy()
    #Menu 
    window2 = Tk()
    menuBar = Menu(window2)
    window2.config(menu=menuBar,bg='pink')
    window2.title("Menu")
    window2.geometry('420x250')
    current_time=""

    Label1=Label(window2,text='SlimFit',font = ("Helvetica 16 bold italic",30),bg = "pink",fg='VioletRed3')
    Label1.grid(column=4,row=2)

    Purchases= Menu(menuBar)
    menuBar.add_cascade(label="Purchases", menu=Purchases)

    Purchases.add_command(label="Products to Buy", command=btn_print)
    Purchases.add_separator()

    Purchases.add_command(label="Buying", command=btn_BuyProduct)
    Purchases.add_separator()

    Purchases.add_command(label="Changing", command=btn_ChangeProduct)
    Purchases.add_separator()

    Purchases.add_command(label="Receipt", command=btn_GetReceipt)
    Purchases.add_separator()

    SavingData = Menu(menuBar)
    menuBar.add_cascade(label="Saving Data", menu=SavingData)

    SavingData.add_command(label="Saving Excel", command=btn_SaveExcel)
    SavingData.add_separator()

    SavingData.add_command(label="Saving SQL", command=btn_SaveDB)
    SavingData.add_separator()

    DataImport = Menu(menuBar)
    menuBar.add_cascade(label="Import Data", menu=DataImport)

    DataImport.add_command(label="Import Excel", command=btn_PrintExcel)
    DataImport.add_separator()

    DataImport.add_command(label="Import SQL", command=btn_PrintDB)
    DataImport.add_separator()
    
    DataImport.add_command(label="Import Text file", command=btn_PrintTxtFile)
    DataImport.add_separator()

    buttonShow=Button(window2,width=12, text="Shopping cart", command=btn_viewMyProduct,bg='white',fg='VioletRed4')
    buttonShow.grid(column = 2,row = 200)
    
    buttonWeather=Button(window2,width=13, text="Current Weather", command=btn_Current_Weather,bg='white',fg='VioletRed4')
    buttonWeather.grid(column =4,row = 200)
    
    buttonTime=Button(window2,width=13, text="Current Time", command= lambda : btn_current_Time(window2),bg='white',fg='VioletRed4')
    buttonTime.grid(column =4,row = 250)

    buttonExit=Button(window2,width=12, text="Exit", command=window2.destroy,bg='white',fg='VioletRed4')
    buttonExit.grid(column =5,row = 200)
    
    Label2=Label(window2, text="Good to see you! "+ customer.name +' : )',font = ("Helvetica 16 italic",15),bg = "pink",fg='VioletRed3')
    Label2.grid(column=4,row=300)
    
    window2.mainloop()


def Welcome():
    window.destroy()
    #Menu Info  
    window1 = Tk()
    window1.title("Info")
    window1.geometry('250x150')
    window1.configure(bg='pink')
    v=IntVar()
    v.set(2)
    lbl1=Label(window1,text='Id',bg = "pink")
    lbl1.grid(column=0,row=0)
    Id= Entry(window1,width=20)
    Id.grid(column=1,row=0) 
    lbl2=Label(window1,text='Name',bg = "pink")
    lbl2.grid(column=0,row=1)
    name= Entry(window1,width=20)
    name.grid(column=1,row=1)  
    lbl3=Label(window1,text='Budget',bg = "pink" )
    lbl3.grid(column=0,row=2)
    money=Entry(window1,width=20)
    money.grid(column=1,row=2)
    Radiobutton(window1, text="Male",bg = "pink", variable=v , value=0,command=lambda:rdb_clicked(v.get())).grid(row=3)
    Radiobutton(window1, text="Female",bg = "pink", variable=v , value=1,command=lambda:rdb_clicked(v.get())).grid(column = 1,row=3)
    button1=Button(window1,width=10, text="Submit", command= lambda:SubmitMsg(window1,Id.get(),name.get(),money.get()),bg='white')
    button1.grid(column = 1,row = 4)
    window1.mainloop()

def rdb_clicked(v):
    customer.gender=v

window = Tk()
window.title("Welcome")
window.geometry('330x380')
window.configure(bg='misty rose')
lbl=Label(window,text="Welcome",font='bold',bg='misty rose',fg='VioletRed3')
lbl.grid(column=0,row=0)
img = ImageTk.PhotoImage(Image.open(r'C:\Users\ellab\OneDrive\שולחן העבודה\Logo.png'))  
Logo = Label(window, image=img,bg='misty rose',width=320)
Logo.grid(column=0,row=5)
Logo.image = img
buttonEntery=Button(window, text="press here to insert Info",font='bold', command=Welcome,bg='pink',fg='VioletRed3',height = 2, width = 18)
buttonEntery.grid(column =0 ,row = 2)




##_______MAIN_______##
def main():
    window.mainloop()

main()



##_______ComboBox______##
import tkinter as tk
from tkinter import ttk

def on_select(e):
    if (e==0) or (e==1) or (e==2):
        print('The branch open between 10:30 to 20:00')
    else:
        print('COMING SOON...')

window = tk.Tk()
window.title('ComboBox')
window.geometry('400x120')

ttk.Label(window, text = "Select the branch :",
        font = ("Helvetica 16 italic",15),background = 'pink', foreground ='VioletRed3').grid(column = 0,
        row = 3)

n = tk.StringVar()
branchchoosen = ttk.Combobox(window, width = 20, textvariable = n)

branchchoosen['values'] = ("Natanya", "Jerusalem","Tel-Aviv","Hadera")
branchchoosen.grid(column = 1, row = 3)
btn = Button(window, text = 'Click me !',font = "Helvetica 16 italic",background = 'pink', foreground ='VioletRed3', command = lambda : on_select(branchchoosen.current()))
btn.grid(row = 6, column = 1, pady = 10, padx = 100)
ttk.Label(window, text = "check hours opening: ",
        font = ("Helvetica 16 italic",10),background = 'VioletRed3', foreground ='pink').grid(column = 0,row = 6)

window.mainloop()

